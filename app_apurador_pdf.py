"""
Apurador de Investimentos (Versão PDF) - Aplicação Web Streamlit

Interface web para apuração de investimentos em promoções usando orçamentos em PDF.
Permite upload da planilha simulador Excel e de PDFs de orçamentos.

Autor: Nivea Project
Data: 2026-02-27
"""

import streamlit as st
import pandas as pd
import io
import os
import tempfile
import pdfplumber
from datetime import datetime
import openpyxl
import openpyxl.styles


# Configuração da página
st.set_page_config(
    page_title="Apurador de Investimentos (PDF)",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para melhorar a aparência
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #2c3e50;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #7f8c8d;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stAlert {
        margin-top: 1rem;
    }
    .upload-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


def validar_colunas_preco_final(df):
    """Valida se a planilha de preço final tem as colunas necessárias"""
    colunas_upper = {str(c).upper().strip(): c for c in df.columns}
    
    col_codigo = None
    for nome in ['CÓDIGO BIZ', 'CODIGO BIZ', 'CODIGO', 'CÓDIGO', 'CÓDIGO SAP']:
        if nome in colunas_upper:
            col_codigo = colunas_upper[nome]
            break
            
    if not col_codigo:
        return False, "A planilha deve conter a coluna 'CÓDIGO BIZ' ou 'CODIGO'"
    
    # Padronizar coluna para CODIGO
    df.rename(columns={col_codigo: 'CODIGO'}, inplace=True)
    return True, "Planilha válida"


def extrair_dados_pdf(arquivo_pdf):
    """Extrai código, quantidade e preço do pdf"""
    linhas_extraidas = []
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        tmp.write(arquivo_pdf.getvalue())
        tmp_path = tmp.name
        
    try:
        # Tenta usar pdfplumber primeiro (para PDFs nativos)
        st.write(f"🔄 Lendo {arquivo_pdf.name}...")
        import fitz
        text_found = False
        doc = fitz.open(tmp_path)
        
        # Iterar sobre as páginas para extrair texto nativo com PyMuPDF (fitz)
        for page in doc:
            words = page.get_text('words')
            if words:
                text_found = True
                # Ordenar palavras por coordenada Y (linha) e depois X (coluna)
                words.sort(key=lambda w: (round(w[1], 1), w[0]))
                
                # Agrupar palavras na mesma linha (tolerância de 5 pontos na vertical)
                lines_dict = {}
                for w in words:
                    y_approx = round(w[1], 0)
                    found = False
                    for k in lines_dict.keys():
                        if abs(k - y_approx) < 5:
                            lines_dict[k].append(w)
                            found = True
                            break
                    if not found:
                        lines_dict[y_approx] = [w]
                
                # Reconstruir o texto linha por linha
                for y in sorted(lines_dict.keys()):
                    line_words = sorted(lines_dict[y], key=lambda w: w[0])
                    linha = ' '.join(w[4] for w in line_words)
                    
                    partes = linha.strip().split()
                    
                    # Usa a mesma heurística de validação das partes
                    if len(partes) >= 6 and partes[0].isdigit():
                        def looks_like_number(s):
                            return any(c.isdigit() for c in s)
                            
                        if len(partes[0]) <= 4 and partes[1].isdigit():
                            codigo = partes[1]
                        else:
                            codigo = partes[0]
                            
                        # Tratamento para casos onde os preços/QTD numéricos se encostam
                        qtd_str = partes[-3]
                        preco_str = partes[-2]
                        
                        try:
                            # Evita linhas de cabeçalho
                            if "total" in linha.lower() or "desconto" in linha.lower() or "pag" in linha.lower():
                                continue
                                
                            if looks_like_number(qtd_str) and looks_like_number(preco_str):
                                linhas_extraidas.append({
                                    'CODIGO': str(codigo).strip(),
                                    'VALOR SKU PAGO': preco_str,
                                    'QUANTIDADE': qtd_str
                                })
                        except:
                            pass
        
        # Se PyMuPDF não achar texto, tenta via OCR com fitz (PyMuPDF) e EasyOCR (para PDFs escaneados)
        if not text_found:
            st.warning("⚠️ PDF parece ser uma imagem/escaneado. Aplicando OCR (isso pode demorar mais)...")
            import fitz
            import easyocr
            import numpy as np
            
            doc = fitz.open(tmp_path)
            reader = easyocr.Reader(['pt'], gpu=False, verbose=False)
            
            for p in range(doc.page_count):
                page = doc[p]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                
                # Converter para numpy array
                if pix.n == 3:
                     img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
                else:
                     img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                     
                resultado_ocr = reader.readtext(img_np)
                
                # Agrupar texto por linhas baseando-se no y_center
                linhas_ocr = []
                for (bbox, text, prob) in resultado_ocr:
                    y_center = (bbox[0][1] + bbox[2][1]) / 2
                    found_line = False
                    for line in linhas_ocr:
                        if abs(line['y'] - y_center) < 15: # tolerância de 15 pixels na vertical
                            line['items'].append((bbox[0][0], text))
                            found_line = True
                            break
                    if not found_line:
                        linhas_ocr.append({'y': y_center, 'items': [(bbox[0][0], text)]})
                
                # Para cada linha do OCR, ordenar e procurar por itens
                for line in linhas_ocr:
                    line['items'].sort(key=lambda x: x[0])
                    text_str = ' '.join([t for x, t in line['items']])
                    partes = text_str.split()
                    
                    if len(partes) >= 5 and partes[0].isdigit():
                        def looks_like_number(s):
                            return any(c.isdigit() for c in s)
                            
                        if len(partes[0]) <= 4 and partes[1].isdigit():
                            codigo = partes[1]
                        else:
                            codigo = partes[0]
                        
                        try:
                            # Pular as linhas de cabeçalho ou totais que por acaso começam com número
                            if "total" in text_str.lower() or "desconto" in text_str.lower(): continue
                            
                            qtd_str = partes[-3]
                            preco_str = partes[-2]
                            
                            if looks_like_number(qtd_str) and looks_like_number(preco_str):
                                linhas_extraidas.append({
                                    'CODIGO': str(codigo).strip(),
                                    'VALOR SKU PAGO': preco_str,
                                    'QUANTIDADE': qtd_str
                                })
                        except:
                            pass
                            
    except Exception as e:
        import traceback
        st.error(f"Erro na extração de texto/OCR: {repr(e)}")
        # não apago tmp_path aqui pra ser limpado no finally, mas open já vai ser safe
    finally:
        try:
            os.unlink(tmp_path)
        except:
            pass
    
    if not linhas_extraidas:
        return False, "Nenhum dado válido de orçamento encontrado no PDF. O formato deve ser parecido com as planilhas da Poliequipes."
        
    df_temp = pd.DataFrame(linhas_extraidas)
    return True, df_temp


def limpar_valor_monetario(valor):
    if pd.isna(valor):
        return None
    
    valor_str = str(valor).strip()
    valor_str = valor_str.replace('R$', '').replace('r$', '').replace('$', '').replace(' ', '').strip()
    
    tem_virgula = ',' in valor_str
    tem_ponto = '.' in valor_str
    
    if tem_virgula and tem_ponto:
        pos_virgula = valor_str.rfind(',')
        pos_ponto = valor_str.rfind('.')
        
        if pos_virgula > pos_ponto:
            valor_str = valor_str.replace('.', '').replace(',', '.')
        else:
            valor_str = valor_str.replace(',', '')
    elif tem_virgula:
        valor_str = valor_str.replace(',', '.')
        
    try:
        return float(valor_str)
    except:
        return None



def processar_dados(df_preco_final, orcamentos_dict):
    """Processa os dados e calcula investimentos"""
    # Copiar dataframe de preço final
    df_resultado = df_preco_final.copy()
    
    # Converter CODIGO para string e remover zeros a esquerda se houver
    df_resultado['CODIGO'] = df_resultado['CODIGO'].astype(str).str.strip().str.lstrip('0')
    
    # Verificar se existe coluna de valor negociado (case-insensitive)
    coluna_valor_negociado = None
    colunas_upper = {str(col).upper().strip(): col for col in df_resultado.columns}
    
    for possivel_nome in ['VALOR NEGOCIADO REDE', 'VALOR NEGOCIADO', 'PRECO NEGOCIADO', 'PREÇO NEGOCIADO']:
        if possivel_nome in colunas_upper:
            coluna_valor_negociado = colunas_upper[possivel_nome]
            break
    
    if coluna_valor_negociado is None:
        st.error("❌ Não foi encontrada coluna de valor negociado na planilha de Preço Final")
        st.info("💡 Colunas aceitas: 'VALOR NEGOCIADO REDE', 'VALOR NEGOCIADO', 'PRECO NEGOCIADO'")
        return None, None
    
    # Converter valor negociado para numérico antes das comparações
    df_resultado[coluna_valor_negociado] = df_resultado[coluna_valor_negociado].apply(limpar_valor_monetario)
    
    # Coletar todos os CODIGOs presentes nos orçamentos
    codigos_orcamentos = set()
    for df_orc in orcamentos_dict.values():
        cods_str = df_orc['CODIGO'].astype(str).str.strip().str.lstrip('0')
        codigos_orcamentos.update(cods_str.tolist())
    
    # Filtrar apenas produtos do Preço Final que estão em algum orçamento
    cods_resultado = df_resultado['CODIGO'].astype(str).str.strip().str.lstrip('0')
    df_no_orcamento = df_resultado[cods_resultado.isin(codigos_orcamentos)]
    
    # Detectar produtos sem preço entre os que estão nos orçamentos
    sem_preco = df_no_orcamento[
        df_no_orcamento[coluna_valor_negociado].isna() | (df_no_orcamento[coluna_valor_negociado] == 0)
    ]
    
    if not sem_preco.empty:
        col_produto = next((c for c in df_resultado.columns if 'produto' in str(c).lower() or 'descri' in str(c).lower()), None)
        col_cod = 'CODIGO' if 'CODIGO' in df_resultado.columns else None
        
        st.error(f"❌ **{len(sem_preco)} produto(s) presentes no orçamento estão sem preço negociado (zero ou vazio)**. Corrija antes de continuar.")
        
        with st.expander("📋 Ver lista de produtos sem preço", expanded=True):
            colunas_exibir = []
            if col_cod:
                colunas_exibir.append(col_cod)
            if col_produto:
                colunas_exibir.append(col_produto)
            colunas_exibir.append(coluna_valor_negociado)
            
            if colunas_exibir:
                st.dataframe(sem_preco[colunas_exibir].reset_index(drop=True), use_container_width=True)
            else:
                st.dataframe(sem_preco.reset_index(drop=True), use_container_width=True)
        
        return None, None
    

    for nome, df_orc in orcamentos_dict.items():
        # Converter CODIGO para string lidando com floats passados s/ doc
        df_orc['CODIGO'] = df_orc['CODIGO'].astype(str).str.strip().str.lstrip('0')
        df_orc['CODIGO'] = df_orc['CODIGO'].str.replace('.0', '', regex=False)
        
        # Preparar dados do orçamento
        df_orc_temp = df_orc[['CODIGO', 'VALOR SKU PAGO', 'QUANTIDADE']].copy()
        
        # Converter valor negociado e numero na base
        df_resultado['CODIGO_TMP'] = df_resultado['CODIGO'].astype(str).str.strip().str.lstrip('0').str.replace('.0', '', regex=False)
        
        # Limpar e converter valores para numérico
        df_orc_temp['VALOR SKU PAGO'] = df_orc_temp['VALOR SKU PAGO'].apply(limpar_valor_monetario)
        df_orc_temp['QUANTIDADE'] = pd.to_numeric(df_orc_temp['QUANTIDADE'].astype(str).str.replace(',', '.'), errors='coerce')
        
        # Agrupar por CODIGO no orçamento caso existam duplicatas no PDF
        df_orc_temp = df_orc_temp.groupby('CODIGO', as_index=False).agg({
            'VALOR SKU PAGO': 'mean', # usar média se houver diferença de preço na msm nota?? Mas de fato soma a quantidade
            'QUANTIDADE': 'sum'
        })
        
        valores_validos = df_orc_temp['VALOR SKU PAGO'].notna().sum()
        qtd_validas = df_orc_temp['QUANTIDADE'].notna().sum()
        st.caption(f"📊 {nome}: {valores_validos} valores SKU válidos, {qtd_validas} quantidades válidas (de {len(df_orc_temp)} códigos únicos)")
        
        df_orc_temp = df_orc_temp.rename(columns={
            'VALOR SKU PAGO': f'{nome}_VALOR_SKU_PAGO',
            'QUANTIDADE': f'{nome}_QUANTIDADE'
        })
        
        # Merge por CODIGO_TMP (LEFT JOIN)
        antes_merge = len(df_resultado)
        df_resultado = df_resultado.merge(
            df_orc_temp,
            left_on='CODIGO_TMP',
            right_on='CODIGO',
            how='left',
            suffixes=('', '_y')
        )
        
        if 'CODIGO_y' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['CODIGO_y'])
            
        # Debug: verificar quantos matches foram feitos
        matches = df_resultado[f'{nome}_QUANTIDADE'].notna().sum()
        st.caption(f"✅ {nome}: {matches} produtos encontrados no Preço Final (de {antes_merge} produtos)")
        
        if matches == 0:
            st.warning(f"⚠️ Nenhum produto de '{nome}' foi encontrado no Preço Final. Verifique se os CÓDIGOS BIZ batem!")
            
    if 'CODIGO_TMP' in df_resultado.columns:
        df_resultado = df_resultado.drop(columns=['CODIGO_TMP'])
    
    # Calcular investimentos e valores para cada orçamento
    for nome in orcamentos_dict.keys():
        
        # Calcular Investimento Total e Valor de Pedido
        df_resultado[f'{nome}_INVESTIMENTO_TOTAL'] = (
            (df_resultado[f'{nome}_VALOR_SKU_PAGO'] - df_resultado[coluna_valor_negociado]) * 
            df_resultado[f'{nome}_QUANTIDADE']
        )
        
        df_resultado[f'{nome}_VALOR_PEDIDO_TOTAL'] = (
            df_resultado[f'{nome}_VALOR_SKU_PAGO'] * df_resultado[f'{nome}_QUANTIDADE']
        )
    
    # Criar colunas totalizadoras (soma de todos os orçamentos)
    colunas_investimento = [f'{nome}_INVESTIMENTO_TOTAL' for nome in orcamentos_dict.keys()]
    colunas_valor_pedido = [f'{nome}_VALOR_PEDIDO_TOTAL' for nome in orcamentos_dict.keys()]
    
    # Somar todas as colunas de investimento (ignorando NaN)
    df_resultado['INVESTIMENTO_TOTAL_GERAL'] = df_resultado[colunas_investimento].sum(axis=1, skipna=True)
    
    # Somar todas as colunas de valor de pedido (ignorando NaN)
    df_resultado['VALOR_PEDIDO_TOTAL_GERAL'] = df_resultado[colunas_valor_pedido].sum(axis=1, skipna=True)
    
    # Coletar estatísticas ANTES de renomear/remover colunas (para exibir no Streamlit)
    estatisticas = {}
    for nome in orcamentos_dict.keys():
        produtos_encontrados = df_resultado[f'{nome}_QUANTIDADE'].notna().sum()
        total_produtos = len(df_resultado)
        estatisticas[nome] = {
            'encontrados': produtos_encontrados,
            'total': total_produtos
        }
    
    # Remover colunas individuais de investimento e valor de pedido (manter só os totais)
    colunas_para_remover = colunas_investimento + colunas_valor_pedido
    df_resultado = df_resultado.drop(columns=colunas_para_remover)
    
    # Renomear colunas para nomes mais amigáveis
    renomeacoes = {}
    
    # Renomear colunas dos orçamentos dinamicamente
    nomes_orcamentos = list(orcamentos_dict.keys())
    for idx, nome in enumerate(nomes_orcamentos, 1):
        renomeacoes[f'{nome}_VALOR_SKU_PAGO'] = f'Preço venda loja {idx}'
        renomeacoes[f'{nome}_QUANTIDADE'] = f'Qtd venda loja {idx}'
    
    # Renomear colunas totais
    renomeacoes['INVESTIMENTO_TOTAL_GERAL'] = 'Verba Total'
    renomeacoes['VALOR_PEDIDO_TOTAL_GERAL'] = 'TT.Pedido'
    
    df_resultado = df_resultado.rename(columns=renomeacoes)
    
    # Calcular % Investimento (Verba Total / TT.Pedido * 100)
    df_resultado['% Investimento'] = (
        (df_resultado['Verba Total'] / df_resultado['TT.Pedido']) * 100
    ).round(2)  # Arredondar para 2 casas decimais
    
    # Substituir inf e NaN por 0 (quando TT.Pedido = 0)
    df_resultado['% Investimento'] = df_resultado['% Investimento'].replace([float('inf'), -float('inf')], 0).fillna(0)
    
    return df_resultado, estatisticas


def converter_df_para_excel(df, nome_rede=""):
    """Converte DataFrame para Excel em memória com formatação e resumo"""
    output = io.BytesIO()
    
    # Calcular totais para o resumo
    total_verba = df['Verba Total'].sum()
    total_pedido = df['TT.Pedido'].sum()
    percentual_investimento = (total_verba / total_pedido * 100) if total_pedido > 0 else 0
    
    # Criar título com nome da rede e data
    data_atual = datetime.now().strftime('%d/%m/%Y')
    titulo = f"RESUMO - {nome_rede} - {data_atual}" if nome_rede else f"RESUMO - {data_atual}"
    
    # Criar writer com engine openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Escrever dados principais a partir da linha 5 (startrow=4)
        df.to_excel(writer, index=False, sheet_name='Apuração', startrow=4)
        
        # Obter worksheet para aplicar formatação
        worksheet = writer.sheets['Apuração']
        
        # Adicionar título do resumo na linha 1
        worksheet['A1'] = titulo
        worksheet['A1'].font = openpyxl.styles.Font(size=11, bold=True)
        
        # Adicionar cabeçalhos do resumo na linha 2
        worksheet['A2'] = 'Verba Total'
        worksheet['B2'] = 'TT.Pedido'
        worksheet['C2'] = '% Investimento'
        
        # Aplicar negrito nos cabeçalhos
        for cell in ['A2', 'B2', 'C2']:
            worksheet[cell].font = openpyxl.styles.Font(bold=True)
        
        # Adicionar valores do resumo na linha 3
        worksheet['A3'] = total_verba
        worksheet['B3'] = total_pedido
        worksheet['C3'] = percentual_investimento
        
        # Formatar valores do resumo
        worksheet['A3'].number_format = 'R$ #,##0.00'
        worksheet['B3'].number_format = 'R$ #,##0.00'
        worksheet['C3'].number_format = '0.00"%"'
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        colunas_moeda = ['Valor Negociado REDE', 'Verba Total', 'TT.Pedido']
        colunas_moeda_dinamicas = [col for col in df.columns if 'Preço venda loja' in col]
        colunas_moeda_dinamicas_valores_reais = [col for col in df.columns if 'VALOR NEGOCIADO' in col.upper()]
        todas_colunas_moeda = colunas_moeda + colunas_moeda_dinamicas + colunas_moeda_dinamicas_valores_reais
        
        coluna_percentual = '% Investimento'
        
        indices_moeda = []
        indice_percentual = None
        indice_codigo = None
        
        for idx, col in enumerate(df.columns, 1):  # Excel columns são 1-indexed
            if col in todas_colunas_moeda or 'VALOR NEGOCIADO' in str(col).upper():
                indices_moeda.append(idx)
            elif col == coluna_percentual:
                indice_percentual = idx
            elif str(col).strip().upper() == 'CODIGO':
                indice_codigo = idx
        
        # Aplicar formatação (começar da linha 6 pois linha 5 é cabeçalho dos dados)
        for row in range(6, len(df) + 6):
            for col_idx in indices_moeda:
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = 'R$ #,##0.00'
            
            if indice_percentual:
                cell = worksheet.cell(row=row, column=indice_percentual)
                cell.number_format = '0.00"%"'
                
            if indice_codigo:
                cell = worksheet.cell(row=row, column=indice_codigo)
                try:
                    # Garantir que o valor seja do tipo int/float em vez de string
                    cell.value = int(cell.value)
                    cell.number_format = '0'
                except:
                    pass
        
        from openpyxl.styles import PatternFill
        
        cor_cabecalho = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid') 
        cor_dados_azul = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid') 
        cor_dados_verde = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') 
        cor_orcamentos = PatternFill(start_color='FEF2CB', end_color='FEF2CB', fill_type='solid') 
        cor_resumo_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
        cor_resumo_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
        cor_resumo_verde = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid') 
        
        for col in range(1, 4): 
            cell = worksheet.cell(row=2, column=col)
            cell.fill = cor_resumo_preto
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        
        worksheet['A3'].fill = cor_resumo_amarelo
        worksheet['B3'].fill = cor_resumo_verde
        worksheet['C3'].fill = cor_resumo_amarelo
        
        total_cols = len(df.columns)
        for col in range(1, total_cols + 1):  
            cell = worksheet.cell(row=5, column=col)
            cell.fill = cor_cabecalho
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        
        for row in range(6, len(df) + 6): 
            for col in range(1, total_cols + 1): 
                cell = worksheet.cell(row=row, column=col)
                if cell.fill.start_color.index == '00000000': # Se não tem preenchimento
                    if col <= 5:
                        cell.fill = cor_dados_azul
        
        
        # Identificar colunas de orçamentos (qtd e preço venda)
        colunas_orcamento = []
        for idx, col_name in enumerate(df.columns, 1):
            col_lower = str(col_name).lower()
            if 'preço venda loja' in col_lower or 'qtd venda loja' in col_lower:
                colunas_orcamento.append(idx)
        
        # Aplicar cor bege claro nas colunas de orçamentos
        for col_idx in colunas_orcamento:
            for row in range(5, len(df) + 6):  
                cell = worksheet.cell(row=row, column=col_idx)
                cell.fill = cor_orcamentos
                if row == 5:  
                    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                    cell.fill = cor_cabecalho
    
    output.seek(0)
    return output.getvalue()


def main():
    """Função principal da aplicação"""
    
    # Cabeçalho
    st.markdown('<div class="main-header">📊 Apurador de Investimentos (Versão PDF)</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Extrai orçamentos de arquivos PDF e gera planilhas consolidadas</div>', unsafe_allow_html=True)
    
    # Botões de ajuda e template
    col_help1, col_help2, col_help3 = st.columns([1, 1, 1])
    
    with col_help1:
        st.link_button(
            "🎥 Tutorial em Vídeo",
            "https://www.loom.com/share/a50e2261f1e84bd6a223160e7b69ad1e",
            help="Assista ao tutorial completo de como usar o sistema",
            use_container_width=True
        )
    
    with col_help2:
        st.link_button(
            "💬 WhatsApp de Suporte",
            "https://wa.me/5534999079685?text=Olá! Preciso de ajuda com o Apurador de Investimentos",
            help="Entre em contato em caso de dúvidas",
            use_container_width=True
        )
    
    with col_help3:
        try:
            template_path = os.path.join(os.path.dirname(__file__), "SIMULADOR_POLIEQUIPES.xlsx")
            with open(template_path, "rb") as file:
                template_data = file.read()
            
            st.download_button(
                "📋 Modelo Padrão Poliequipes",
                data=template_data,
                file_name="SIMULADOR_POLIEQUIPES.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Baixe a planilha modelo padrão Poliequipes",
                use_container_width=True
            )
        except FileNotFoundError:
            st.warning("⚠️ Planilha modelo Poliequipes não encontrada")
    
    st.markdown("---")
    
    # Barra lateral com instruções
    with st.sidebar:
        st.header("📋 Instruções")
        st.markdown("""
        ### Como usar (Modo PDF):
        
        1️⃣ **Preço Final (Excel)**  
        Upload da planilha Simulador com:
        - CÓDIGO BIZ ou CÓDIGO
        - Valor Negociado REDE
        
        2️⃣ **Orçamentos (PDF)**  
        Upload de 1 ou mais arquivos **PDF** de Orçamento da Poliequipes
        
        3️⃣ **Processar**  
        Clique no botão para consolidar
        
        4️⃣ **Download**  
        Baixe o resultado final em Excel
        """)
        
    # Inicializar session state
    if 'df_preco_final' not in st.session_state:
        st.session_state.df_preco_final = None
    if 'orcamentos_dict' not in st.session_state:
        st.session_state.orcamentos_dict = {}
    if 'df_resultado' not in st.session_state:
        st.session_state.df_resultado = None
    
    # Container para upload de Preço Final
    st.subheader("1️⃣ Planilha Simulador (Preço Final)")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        arquivo_preco = st.file_uploader(
            "Selecione a planilha Simulador (Excel)",
            type=['xlsx', 'xls'],
            key="upload_preco",
            help="Upload do arquivo Excel contendo CÓDIGO BIZ e Preço Negociado"
        )
    
    with col2:
        nome_rede_input = st.text_input(
            "Nome da Rede",
            placeholder="Ex: REDE ABC",
            help="Nome da rede para incluir no relatório"
        )
    
    if arquivo_preco:
        try:
            df_preco = pd.read_excel(arquivo_preco)
            valido, mensagem = validar_colunas_preco_final(df_preco)
            
            if valido:
                st.session_state.df_preco_final = df_preco
                st.session_state.nome_rede = nome_rede_input if nome_rede_input else "[REDE]"
                st.success(f"✅ Simulador carregado com sucesso!")
                with col2:
                    st.metric("📦 Produtos", len(df_preco))
                
                with st.expander("👁️ Visualizar dados carregados"):
                    st.dataframe(df_preco.head(10), use_container_width=True)
            else:
                st.error(f"❌ {mensagem}")
                st.session_state.df_preco_final = None
                
        except Exception as e:
            st.error(f"❌ Erro ao ler arquivo de Excel: {str(e)}")
            st.session_state.df_preco_final = None
    
    st.divider()
    
    # Container para upload de Orçamentos PDF
    st.subheader("2️⃣ Orçamentos em PDF")
    
    st.info("💡 Carregue os PDFs gerados pelo sistema com orçamentos.")
    
    arquivos_orcamento = st.file_uploader(
        "Selecione um ou mais PDFs de orçamento",
        type=['pdf'],
        accept_multiple_files=True,
        key="upload_orcamentos_pdf"
    )
    
    if arquivos_orcamento:
        st.session_state.orcamentos_dict = {}
        
        for arquivo in arquivos_orcamento:
            try:
                valido, resultado = extrair_dados_pdf(arquivo)
                
                if valido:
                    df_temp = resultado
                    nome_orcamento = arquivo.name.replace('.pdf', '')
                    st.session_state.orcamentos_dict[nome_orcamento] = df_temp
                    
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        st.success(f"✅ {arquivo.name}")
                    with col2:
                        st.metric("📦 Linhas Extraídas", len(df_temp))
                    with col3:
                        with st.expander("👁️ Ver Dados"):
                            st.dataframe(df_temp.head(5), use_container_width=True)
                else:
                    st.error(f"❌ {arquivo.name}: {resultado}")
                    
            except Exception as e:
                import traceback
                st.error(f"❌ Erro ao ler {arquivo.name}: {str(e)}")
                st.caption(traceback.format_exc())
    
    st.divider()
    
    # Botão de processamento
    st.subheader("3️⃣ Processar Dados")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        processar_btn = st.button(
            "🚀 Processar e Calcular Investimentos",
            type="primary",
            use_container_width=True,
            disabled=(st.session_state.df_preco_final is None or not st.session_state.orcamentos_dict)
        )
    
    if processar_btn:
        with st.spinner("⏳ Cruzando e calculando dados..."):
            resultado = processar_dados(
                st.session_state.df_preco_final,
                st.session_state.orcamentos_dict
            )
            
            if resultado is None or resultado[0] is None:
                pass 
            else:
                df_resultado, estatisticas = resultado
                st.session_state.df_resultado = df_resultado
                
                st.success("✅ Processamento concluído com sucesso!")
                
                st.subheader("📈 Resumo do Processamento")
                
                cols = st.columns(len(estatisticas) + 1)
                
                with cols[0]:
                    st.metric(
                        "Total de Produtos",
                        len(df_resultado),
                        help="Total de produtos contidos no Simulador Excel"
                    )
                
                for idx, (nome, stats) in enumerate(estatisticas.items(), 1):
                    with cols[idx]:
                        st.metric(
                            f"📦 {nome}",
                            f"{stats['encontrados']}/{stats['total']}",
                            help=f"Produtos validados com o Simulador"
                        )
    
    if st.session_state.df_resultado is not None:
        st.divider()
        st.subheader("4️⃣ Resultado")
        
        with st.expander("👁️ Visualizar Resultado Completo", expanded=False):
            st.dataframe(st.session_state.df_resultado, use_container_width=True)
        
        nome_arquivo = f"Apuracao_PDF_Investimento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_data = converter_df_para_excel(
            st.session_state.df_resultado,
            st.session_state.get('nome_rede', '')
        )
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label="📥 Download Resultado (Excel)",
                data=excel_data,
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        st.info("💡 O arquivo Excel contém todos os dados do Simulador mais as colunas " +
                "de Orçamento vindas dos PDFs processados!")
    
    st.divider()
    st.markdown(
        "<div style='text-align: center; color: #7f8c8d; font-size: 0.9rem;'>"
        "Apurador de Investimentos (PDF) | Streamlit Web App | 2026"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
